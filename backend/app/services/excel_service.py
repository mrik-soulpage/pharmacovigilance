"""
Excel export service for generating tracker reports
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import List
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)


class ExcelService:
    """Service for generating Excel tracker reports"""
    
    def generate_tracker(
        self,
        search_results: List[dict],
        output_path: str,
        week_number: str = "XX"
    ) -> str:
        """
        Generate Excel tracker from search results
        
        Args:
            search_results: List of search result dictionaries with article and product info
            output_path: Path to save the Excel file
            week_number: Week number for the report
            
        Returns:
            Path to generated Excel file
        """
        try:
            logger.info(f"Generating Excel tracker with {len(search_results)} results")
            
            # Prepare data for DataFrame
            tracker_data = []
            
            for result in search_results:
                article = result.get("article", {})
                product = result.get("product", {})
                
                # Format ICSR status
                if result.get("is_icsr") is True:
                    icsr_status = "Y"
                elif result.get("is_icsr") is False:
                    icsr_status = "N"
                else:
                    icsr_status = "NA"
                
                # Format ownership exclusion
                if result.get("ownership_excluded") is True:
                    ownership_excluded = "Y"
                elif result.get("ownership_excluded") is False:
                    ownership_excluded = "N"
                else:
                    ownership_excluded = "NA" if icsr_status == "Y" else ""
                
                # Format other safety info
                if result.get("other_safety_info") is True:
                    other_safety = "Y"
                elif result.get("other_safety_info") is False:
                    other_safety = "N"
                else:
                    other_safety = "NA"
                
                row = {
                    "INN": product.get("inn", ""),
                    "Date of search": result.get("search_date", datetime.now().strftime("%Y-%m-%d")),
                    "Search period (From)": result.get("date_from", ""),
                    "Search period (To)": result.get("date_to", ""),
                    "Search strategy": product.get("search_strategy", ""),
                    "Number of results": 1,
                    "PMID*": article.get("pmid", ""),
                    "Title*": article.get("title", ""),
                    "Authors*": article.get("authors", ""),
                    "Citation*": article.get("citation", ""),
                    "First Author*": article.get("first_author", ""),
                    "Journal/ Book*": article.get("journal", ""),
                    "Publication Year*": article.get("publication_year", ""),
                    "Create Date*": article.get("create_date", ""),
                    "PMCID*": article.get("pmcid", ""),
                    "NIHMS ID*": article.get("nihms_id", ""),
                    "DOI*": article.get("doi", ""),
                    "ICSR (Y/N/NA)": icsr_status,
                    "ICSR description (if applicable)": result.get("icsr_description", ""),
                    "Hikma ownership can be excluded (Y/N/NA)": ownership_excluded,
                    "Reason for exclusion (if applicable)": result.get("exclusion_reason", ""),
                    "ICSR is a duplicate (Y/N/NA)": "Y" if result.get("is_duplicate") else ("N" if icsr_status == "Y" else ""),
                    "Minimum criteria for reporting available? (Y/N/NA)": "Y" if result.get("minimum_criteria_available") else ("N" if icsr_status == "Y" else ""),
                    "Full article available (Y/N/NA)": "Y" if article.get("full_text_available") else ("N" if icsr_status == "Y" else ""),
                    "Date reference sent to PV Operations (if applicable)": result.get("date_sent_to_provider", ""),
                    "Date article ordered (if applicable)": "",
                    "Date article sent to PV Operations (if applicable)": "",
                    "ICSR code (if applicable)": result.get("icsr_code", ""),
                    "Other safety information (Y/N/NA)": other_safety,
                    "Justification for answer in column AC": result.get("safety_info_justification", ""),
                    "Conducted by": result.get("reviewed_by", "AI System"),
                    "Qc'd by": result.get("qc_by", ""),
                    "Comments": result.get("comments", ""),
                }
                
                tracker_data.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(tracker_data)
            
            # Create Excel file with multiple sheets
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                # Write main tracker sheet
                df.to_excel(writer, sheet_name=f"Week {week_number}", index=False)
                
                # Write legends sheet
                self._create_legends_sheet(writer)
            
            # Apply formatting
            self._apply_formatting(output_path, f"Week {week_number}")
            
            logger.info(f"Excel tracker generated successfully: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error generating Excel tracker: {str(e)}")
            raise
    
    def _create_legends_sheet(self, writer):
        """Create legends sheet with column descriptions"""
        legends_data = {
            "Column": [
                "INN",
                "Date of search",
                "Search period (From/To)",
                "Search strategy",
                "Number of results",
                "PMID*",
                "Title*",
                "Authors*",
                "Citation*",
                "First Author*",
                "Journal/ Book*",
                "Publication Year*",
                "Create Date*",
                "PMCID*",
                "NIHMS ID*",
                "DOI*",
                "ICSR (Y/N/NA)",
                "ICSR description",
                "Hikma ownership can be excluded",
                "Reason for exclusion",
                "ICSR is a duplicate",
                "Minimum criteria for reporting available?",
                "Full article available",
                "Date reference sent to PV Operations",
                "Date article ordered",
                "Date article sent to PV Operations",
                "ICSR code",
                "Other safety information",
                "Justification",
                "Conducted by",
                "Qc'd by",
                "Comments"
            ],
            "Description": [
                "International Nonproprietary Name - Generic drug name",
                "Date when the search was conducted",
                "Date range for the literature search",
                "Boolean search query used in PubMed",
                "Number of articles found in the search",
                "PubMed Identifier - Unique article ID",
                "Article title",
                "List of article authors",
                "Full citation for the article",
                "First author name",
                "Journal or book name where article was published",
                "Year of publication",
                "Date when article was added to PubMed database",
                "PubMed Central Identifier",
                "National Institute of Health Manuscript Submission Identifier",
                "Digital Object Identifier",
                "Whether article contains an Individual Case Safety Report (Y=Yes, N=No, NA=No results)",
                "Description of the identified ICSR and adverse events",
                "Whether Hikma's ownership can be excluded for this ICSR (Y/N/NA)",
                "Reasons for excluding Hikma's ownership (territory, brand, dosage form, etc.)",
                "Whether this ICSR was previously identified (Y/N/NA)",
                "Whether the four minimum criteria for reporting are available (Y/N/NA)",
                "Whether full article text is available without additional fees (Y/N/NA)",
                "Date when reference was sent to PV Operations",
                "Date when article was ordered for full-text review",
                "Date when full article was sent to PV Operations",
                "Code received from third-party service provider for validated ICSR",
                "Whether article contains valuable safety information for other activities (Y/N/NA)",
                "Explanation for safety information classification",
                "Name of team member who conducted the search and evaluation",
                "Name of team member who performed quality check",
                "Any additional comments or notes"
            ]
        }
        
        legends_df = pd.DataFrame(legends_data)
        legends_df.to_excel(writer, sheet_name="Legends", index=False)
    
    def _apply_formatting(self, file_path: str, sheet_name: str):
        """Apply formatting to Excel file"""
        try:
            workbook = load_workbook(file_path)
            worksheet = workbook[sheet_name]
            
            # Header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            worksheet.freeze_panes = "A2"
            
            workbook.save(file_path)
            logger.info("Formatting applied successfully")
            
        except Exception as e:
            logger.error(f"Error applying formatting: {str(e)}")

